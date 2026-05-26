import io
import os
from copy import deepcopy
from typing import Dict, List

from .base import FileProcessor


class OpenPyxlProcessor(FileProcessor):
    """
    Unified Excel file processing class, supports in-memory file processing
    """

    def process_file(self, file_data: bytes, chunking_strategy: str, filename: str, **params) -> List[Dict]:
        """Process Excel file in memory"""
        return self._process_excel(
            file_data=file_data, chunking_strategy=chunking_strategy, filename=filename, **params
        )

    def _process_excel(
        self, file_data: bytes, chunking_strategy: str = "basic", filename: str = "", **params
    ) -> List[Dict]:
        """
        Core Excel processing logic, supports byte data input
        """
        # Load workbook
        wb_original, wb_copy = self._load_workbook(file_data)

        # Extract content
        raw_content = self._extract_content(wb_original, wb_copy)

        # Convert to standardized chunk format
        chunks = self._convert_to_chunks(raw_content, filename)

        return chunks

    def _load_workbook(self, file_data: bytes):
        """Load Excel workbook"""
        import openpyxl

        try:
            file_obj = io.BytesIO(file_data)
            wb_original = openpyxl.load_workbook(file_obj)

            wb_copy = deepcopy(wb_original)
            return wb_original, wb_copy

        except Exception as e:
            raise Exception(f"Failed to load Excel file: {str(e)}")

    def _extract_content(self, wb_original, wb_copy) -> List[str]:
        """Extract content from all worksheets"""
        contents = []

        for sheet_name in wb_original.sheetnames:
            sheet = wb_original[sheet_name]
            sheet_copy = wb_copy[sheet_name]

            if self._is_single_column(sheet):
                # Process single column data
                content = self._process_single_column(sheet, sheet_name)
            else:
                # Process multi-column table data
                content = self._process_multi_column(sheet, sheet_copy, sheet_name)

            contents.extend(content)

        return contents

    def _convert_to_chunks(self, raw_content: List[str], filename: str) -> List[Dict]:
        """Convert raw content to standardized chunk format"""
        chunks = []

        for i, content_text in enumerate(raw_content):
            # Determine file type
            file_type = self._determine_file_type(filename)

            chunk = {
                "content": content_text,
                "filename": filename,
                "metadata": {"chunk_index": i, "file_type": file_type},
            }
            chunks.append(chunk)

        return chunks

    def _determine_file_type(self, filename: str) -> str:
        """Determine Excel file type"""
        if filename and filename.lower().endswith(".xlsx"):
            return "xlsx"
        else:
            return "xls"

    def _is_single_column(self, sheet) -> bool:
        """Check if it's single column data"""
        _, max_col = self._get_title_row(sheet)
        return max_col < 2

    def _process_single_column(self, sheet, sheet_name: str) -> List[str]:
        """Process single column data"""
        content_str = ""

        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                # Process first non-empty cell
                cell_value = next((cell for cell in row if cell is not None), "")
                content_str += str(cell_value).replace("\n", "<br>") + "\n"

        return [content_str + "\n————" + sheet_name]

    def _process_multi_column(self, sheet, sheet_copy, sheet_name: str) -> List[str]:
        """Process multi-column table data"""
        # Get title row position
        begin_row, _ = self._get_title_row(sheet_copy)

        # Process merged cells
        self._merge_all_cells(sheet_copy)

        # Get title and remarks
        title_key = self._get_title_key(begin_row, sheet)
        remark = self._get_remark(sheet, begin_row)

        # Extract table content
        content = self._extract_table_content(title_key, remark, sheet, begin_row, sheet_name)

        return content

    def _get_title_row(self, sheet) -> tuple:
        """Get title row position and maximum column count"""
        max_col = 0
        position_max_col = 0

        # First process column merging
        self._merge_columns(sheet)

        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            non_empty_cells = sum(1 for cell in row if cell.value is not None)
            if non_empty_cells > max_col:
                max_col = non_empty_cells
                position_max_col = row_idx

        return position_max_col, max_col

    def _get_remark(self, sheet, begin_row: int) -> str:
        """Get remarks before the title"""
        remark = ""

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None for cell in row):
                continue
            if row_idx >= begin_row:
                break
            remark += "<br>" + self._join_tuple_elements(row)

        return remark

    def _get_title_key(self, begin_row: int, sheet) -> List[str]:
        """Get column headers from title row"""
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None for cell in row):
                continue
            if row_idx == begin_row:
                return [str(cell) if cell is not None else "" for cell in row]
        return []

    def _extract_table_content(
        self, title_key: List[str], remark: str, sheet, begin_row: int, sheet_name: str
    ) -> List[str]:
        """Extract table content and convert to markdown format"""
        content = []

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None for cell in row):
                continue
            if row_idx <= begin_row:
                continue

            # Build current row content
            row_content = self._build_row_content(title_key, row, remark, sheet_name)
            content.append(row_content)

        return content

    def _build_row_content(self, title_key: List[str], row: tuple, remark: str, sheet_name: str) -> str:
        """Build single row content"""
        # Add remark column
        if remark:
            title_key_with_remark = title_key + ["Remark before title"]
            row_with_remark = row + (remark,)
        else:
            title_key_with_remark = title_key
            row_with_remark = row

        # Build key-value pair dictionary
        result = {}
        for k, v in zip(title_key_with_remark, row_with_remark):
            key = str(k).replace("\n", "<br>") if k is not None else ""
            value = str(v).replace("\n", "<br>") if v is not None else ""
            result[key] = value

        # Convert to markdown table
        markdown_table = self._dict_to_markdown_table(result)
        return markdown_table + "\n————" + sheet_name

    def _merge_columns(self, sheet):
        """Process column merging"""
        merged_ranges = list(sheet.merged_cells.ranges)

        # Unmerge cells
        for merged_range in merged_ranges:
            if str(merged_range)[0] != str(merged_range)[3]:
                continue
            sheet.unmerge_cells(str(merged_range))

        # Fill merged area values
        for merged_range in merged_ranges:
            if str(merged_range)[0] != str(merged_range)[3]:
                continue
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = sheet.cell(row=min_row, column=min_col).value

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value

    def _merge_all_cells(self, sheet):
        """Process all merged cells"""
        merged_ranges = list(sheet.merged_cells.ranges)

        # Unmerge all cells
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))

        # Fill all merged area values
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = sheet.cell(row=min_row, column=min_col).value

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value

    @staticmethod
    def _dict_to_markdown_table(data: Dict[str, str]) -> str:
        """Convert dictionary to markdown table"""
        if not data:
            return ""

        keys = []
        values = []

        for key, value in data.items():
            keys.append("None" if key is None else str(key))
            values.append("None" if value is None else str(value))

        # Build markdown table
        table = "| " + " | ".join(keys) + " |\n"
        table += "| " + " | ".join(["---"] * len(keys)) + " |\n"
        table += "| " + " | ".join(values) + " |"

        return table

    @staticmethod
    def _join_tuple_elements(input_tuple: tuple) -> str:
        """Join non-empty elements in tuple"""
        return ";".join(str(item) for item in input_tuple if item is not None)

    @staticmethod
    def _check_file_exists(file_path: str) -> bool:
        """Check if file exists and is readable"""
        return os.path.isfile(file_path) and os.access(file_path, os.R_OK)
