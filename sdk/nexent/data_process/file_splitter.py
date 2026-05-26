import csv
import json
import math
import os
import subprocess
import tempfile
import xml.etree.ElementTree as ET
from copy import copy
from io import BytesIO, StringIO, TextIOWrapper
from typing import List


class FileSplitter:

    def split_csv_by_size(self, csv_bytes, max_size, encoding="utf-8"):
        text = csv_bytes.decode(encoding)
        reader = list(csv.reader(StringIO(text)))

        if not reader:
            return []

        header = reader[0]
        rows = reader[1:]

        result = []

        def build_csv_bytes(sub_rows):
            buffer = StringIO()
            writer = csv.writer(buffer)

            writer.writerow(header)
            writer.writerows(sub_rows)

            return buffer.getvalue().encode(encoding)

        def split_range(start, end):
            sub_rows = rows[start:end]
            csv_part = build_csv_bytes(sub_rows)

            size = len(csv_part)
            row_count = end - start

            if size <= max_size or row_count <= 1:
                result.append(BytesIO(csv_part))
                return

            group_count = math.ceil(size / max_size)
            group_count = min(group_count, row_count)
            rows_per_group = math.ceil(row_count / group_count)

            current = start
            for _ in range(group_count):
                next_end = min(current + rows_per_group, end)
                if current >= next_end:
                    break

                split_range(current, next_end)
                current = next_end

        split_range(0, len(rows))

        return result

    def split_epub_by_size(self, epub_bytes, max_size):
        import ebooklib
        from ebooklib import epub

        book = epub.read_epub(BytesIO(epub_bytes))
        items = list(book.get_items_of_type(ebooklib.ITEM_DOCUMENT))

        result: List[BytesIO] = []

        def build_epub(parts):
            new_book = epub.EpubBook()

            new_book.set_title(
                book.get_metadata("DC", "title")[0][0]
                if book.get_metadata("DC", "title")
                else "split"
            )

            new_items = []

            for i, item in enumerate(parts):
                new_item = epub.EpubHtml(
                    title=item.get_name(),
                    file_name=f"chap_{i}.xhtml",
                    content=item.get_content(),
                )
                new_book.add_item(new_item)
                new_items.append(new_item)

            new_book.toc = tuple(new_items)
            new_book.spine = new_items

            buffer = BytesIO()
            epub.write_epub(buffer, new_book)
            return buffer.getvalue()

        def split_chunks(chapters):
            epub_part = build_epub(chapters)
            size = len(epub_part)

            if size <= max_size or len(chapters) <= 1:
                result.append(BytesIO(epub_part))
                return

            group_count = math.ceil(size / max_size)
            group_count = min(group_count, len(chapters))
            per_group = math.ceil(len(chapters) / group_count)

            for i in range(0, len(chapters), per_group):
                sub = chapters[i : i + per_group]
                split_chunks(sub)

        split_chunks(items)

        return result


    def copy_images_safe(self, src_ws, dst_ws):
        from openpyxl.drawing.image import Image

        if not hasattr(src_ws, "_images") or not src_ws._images:
            return

        for img in src_ws._images:
            try:
                img_bytes = None

                if hasattr(img, "_data"):
                    try:
                        img_bytes = img._data()
                    except Exception:
                        img_bytes = None

                if img_bytes is None:
                    continue

                bio = BytesIO(img_bytes)
                new_img = Image(bio)

                try:
                    anchor = copy(img.anchor)
                except Exception:
                    anchor = img.anchor

                dst_ws.add_image(new_img, anchor)

            except Exception:
                continue

    def split_excel(self, excel_bytes, max_size):
        from openpyxl import Workbook, load_workbook

        file_size = len(excel_bytes)

        if file_size <= max_size:
            return [BytesIO(excel_bytes)]

        wb = load_workbook(BytesIO(excel_bytes), data_only=False)

        sheet_data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            header = rows[0]
            data = rows[1:] if len(rows) > 1 else []

            if not data and all(v is None for v in header):
                continue

            sheet_data[sheet_name] = {
                "header": header,
                "data": data,
                "src_ws": ws,
            }

        if not sheet_data:
            return []

        group_count = math.ceil(file_size / max_size)

        results = []

        for g in range(group_count):
            new_wb = Workbook()
            new_wb.remove(new_wb.active)

            has_data = False

            for sheet_name, content in sheet_data.items():
                header = content["header"]
                data = content["data"]
                src_ws = content["src_ws"]

                chunk_size = math.ceil(len(data) / group_count) if data else 0

                start = g * chunk_size
                end = start + chunk_size

                chunk = data[start:end]

                if not chunk:
                    continue

                ws = new_wb.create_sheet(title=sheet_name)
                ws.append(list(header))

                for row in chunk:
                    ws.append(list(row) if row else [])

                self.copy_images_safe(src_ws, ws)

                has_data = True

            if not has_data:
                continue

            buffer = BytesIO()
            new_wb.save(buffer)

            results.append(BytesIO(buffer.getvalue()))

        return results


    def split_json_stream(self, json_bytes, max_size):
        import ijson

        buffer = BytesIO(json_bytes)
        items = ijson.items(buffer, "item")

        result: List[BytesIO] = []
        batch = []
        current_size = 0

        for item in items:
            item_bytes = json.dumps(item, ensure_ascii=False).encode("utf-8")
            if current_size + len(item_bytes) > max_size and batch:
                result.append(BytesIO(self._json_bytes_from_batch(batch)))
                batch = []
                current_size = 0

            batch.append(item)
            current_size += len(item_bytes)

        if batch:
            result.append(BytesIO(self._json_bytes_from_batch(batch)))

        return result


    def _json_bytes_from_batch(self, data):
        return json.dumps(data, ensure_ascii=False).encode("utf-8")

    def split_markdown(self, md_bytes, max_size):
        text = md_bytes.decode("utf-8")
        result = []

        def find_highest_header_level(content):
            for level in range(1, 7):
                header_mark = "#" * level + " "
                if header_mark in content:
                    return level
            return 1

        def split_by_level(content, level, parent_headers):
            from langchain_text_splitters import MarkdownHeaderTextSplitter
            if len(content.encode("utf-8")) <= max_size or level > 6:
                result.append(BytesIO(content.encode("utf-8")))
                return
            
            headers_to_split_on = [(f"{'#' * level}", f"h{level}")]
            splitter = MarkdownHeaderTextSplitter(headers_to_split_on=headers_to_split_on)
            docs = splitter.split_text(content)

            if len(docs) <= 1:
                split_by_level(content, level + 1, parent_headers)
                return

            for doc in docs:
                chunk = doc.page_content
                current_header = doc.metadata.get(f"h{level}", "")

                full_headers = parent_headers.copy()
                if current_header:
                    full_headers.append((level, current_header))

                header_text = ""
                for lvl, h in full_headers:
                    header_text += f"{'#' * lvl} {h}\n"

                new_content = header_text + chunk
                split_by_level(new_content, level + 1, full_headers)

        start_level = find_highest_header_level(text)
        split_by_level(text, start_level, [])

        return result


    def split_pdf_by_size(self, pdf_bytes, max_size):
        from pypdf import PdfReader, PdfWriter

        reader = PdfReader(BytesIO(pdf_bytes))
        total_pages = len(reader.pages)

        result = []

        def build_pdf_bytes(start, end):
            writer = PdfWriter()
            for i in range(start, end):
                writer.add_page(reader.pages[i])

            buffer = BytesIO()
            writer.write(buffer)
            return buffer.getvalue()

        def split_range(start, end):
            pdf_part = build_pdf_bytes(start, end)
            size = len(pdf_part)
            page_count = end - start

            if size <= max_size or page_count <= 1:
                result.append(BytesIO(pdf_part))
                return

            group_count = math.ceil(size / max_size)
            group_count = min(group_count, page_count)
            pages_per_group = math.ceil(page_count / group_count)

            current = start
            for _ in range(group_count):
                next_end = min(current + pages_per_group, end)
                if current >= next_end:
                    break

                split_range(current, next_end)
                current = next_end

        split_range(0, total_pages)

        return result


    def split_txt_by_size(self, txt_bytes, max_size, encoding="utf-8"):
        buffer = BytesIO(txt_bytes)
        reader = TextIOWrapper(buffer, encoding=encoding)

        result: List[BytesIO] = []
        current_size = 0
        current_lines = []

        def flush_part(lines):
            text = "".join(lines)
            part_bytes = text.encode(encoding)
            result.append(BytesIO(part_bytes))

        for line in reader:
            line_size = len(line.encode(encoding))

            if current_size + line_size > max_size and current_size > 0:
                flush_part(current_lines)
                current_lines = []
                current_size = 0

            current_lines.append(line)
            current_size += line_size

        if current_lines:
            flush_part(current_lines)

        reader.close()

        return result


    def split_xml_by_size(self, xml_bytes, max_size):
        root = ET.fromstring(xml_bytes)
        children = list(root)

        result: List[BytesIO] = []

        def build_xml_bytes(elements):
            new_root = ET.Element(root.tag, root.attrib)

            for elem in elements:
                new_root.append(elem)

            return ET.tostring(new_root, encoding="utf-8")

        def split_range(elements):
            xml_part = build_xml_bytes(elements)
            size = len(xml_part)

            if size <= max_size or len(elements) <= 1:
                result.append(BytesIO(xml_part))
                return

            group_count = math.ceil(size / max_size)
            group_count = min(group_count, len(elements))
            per_group = math.ceil(len(elements) / group_count)

            for i in range(0, len(elements), per_group):
                sub = elements[i : i + per_group]
                split_range(sub)

        split_range(children)

        return result


    def _convert_bytes_with_libreoffice(
        self, input_bytes, input_ext, output_ext, libreoffice_path="soffice"
    ):
        with tempfile.TemporaryDirectory() as tmpdir:
            src_path = os.path.join(tmpdir, f"input{input_ext}")
            with open(src_path, "wb") as f:
                f.write(input_bytes)

            cmd = [
                libreoffice_path,
                "--headless",
                "--convert-to",
                output_ext.lstrip("."),
                "--outdir",
                tmpdir,
                src_path,
            ]

            try:
                subprocess.run(
                    cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE
                )
            except Exception as exc:
                raise RuntimeError(f"LibreOffice conversion failed: {exc}")

            output_path = os.path.join(tmpdir, f"input{output_ext}")
            if not os.path.exists(output_path):
                candidates = [
                    f
                    for f in os.listdir(tmpdir)
                    if f.lower().endswith(output_ext.lower())
                ]
                if not candidates:
                    raise RuntimeError("LibreOffice conversion produced no output")
                output_path = os.path.join(tmpdir, candidates[0])

            with open(output_path, "rb") as f:
                return f.read()

    def file_process(self, file_data, filename, max_size, **kwargs) -> List[BytesIO]:
        ext = os.path.splitext(filename)[1].lower()

        if ext in {".doc", ".docx"}:
            libreoffice_path = kwargs.get("libreoffice_path", "soffice")
            pdf_bytes = self._convert_bytes_with_libreoffice(
                file_data, ext, ".pdf", libreoffice_path=libreoffice_path
            )
            pdf_parts = self.split_pdf_by_size(pdf_bytes, max_size=max_size)

            # If no actual split happened, keep original Word bytes as-is.
            if not pdf_parts or len(pdf_parts) == 1:
                return [BytesIO(file_data)]

            # For real splits, keep PDF parts and let downstream parsing use PDF bytes
            # while filenames remain as Word (handled by caller).
            return pdf_parts

        if ext == ".csv":
            return self.split_csv_by_size(
                file_data,
                max_size=max_size,
                encoding=kwargs.get("encoding", "utf-8"),
            )

        if ext == ".epub":
            return self.split_epub_by_size(file_data, max_size=max_size)

        if ext in {".xlsx", ".xls"}:
            return self.split_excel(file_data, max_size=max_size)

        if ext == ".json":
            return self.split_json_stream(file_data, max_size=max_size)

        if ext == ".md":
            return self.split_markdown(file_data, max_size=max_size)

        if ext == ".pdf":
            return self.split_pdf_by_size(file_data, max_size=max_size)

        if ext == ".txt":
            return self.split_txt_by_size(
                file_data,
                max_size=max_size,
                encoding=kwargs.get("encoding", "utf-8"),
            )

        if ext == ".xml":
            return self.split_xml_by_size(file_data, max_size=max_size)

        raise ValueError(f"Unsupported file extension: {ext}")
